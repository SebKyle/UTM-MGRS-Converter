# Created by Jochuah Kyle
#Converts between UTM and MGRS for an Army Mission Commmand System, works for a specific format, but also has tools to convert with user input
#code can be adjusted to work for other system, it is specifically formatted to work with an import function on the AFATDS (Advanced Field Artillery Tactical Data System)
# before using do pip install --upgrade pip, pip install openpyxl, pip install mgrs if you have not already
import openpyxl
import mgrs
from openpyxl.utils import column_index_from_string

#in order to change the workbook change the filepath below 
#YOU HAVE TO USE \\ INSTEAD OF / WHENEVER SELECTING FILE PATH OR IT WILL NOT WORK
#don't delete the # symbols as that will also break how the code works
workbook = openpyxl.load_workbook('C:\\Users\\Jochuah Kyle\\Desktop\\Coding Stuff\\UTM to MGRS Python Converter\\UTM TEST FILE.xlsx')
columnLetter = column_index_from_string('G')
def userInputConverterUTMToMGRS():
    convert = mgrs.MGRS()
    utm_zone = int(input("Input the UTM zone: "))
    utm_easting = float(input("Input the UTM Easting: "))
    utm_northing = float(input("Input the UTM Northing: "))
    utm_hemisphere = input("Input the UTM Hemisphere(N for north S for south): ")

    #converting UTM to MGRS
    mgrs_coordinates = convert.UTMToMGRS(utm_zone,utm_hemisphere,utm_easting,utm_northing)
    print(f"MGRS Coordinates: {mgrs_coordinates}")

def userInputConverterMGRSToUTM():
    mgrs_coordinates= input("Input the MGRS Coordinates: ")
    convert = mgrs.MGRS()
    UTM_coordinates = convert.MGRSToUTM(mgrs_coordinates)
    
    print(UTM_coordinates)
# Automatic conversion from UTM to MGRS
def UTMToMGRSConverter():
    utm_hemisphere = input("Input the UTM Hemisphere(N for north S for south): ")
    convert = mgrs.MGRS()
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]

        print(f"Changing Sheet: {sheet_name}")

        #Getting all of the from columns and coordinating them
        for row in sheet.iter_rows(min_row=2,min_col=columnLetter, max_col=columnLetter):
            for cell in row:
                #splitting values in order to have useable format for the MGRS converter
                cellValue = cell.value
                splitValues=cellValue.split()
                utm_easting = float(splitValues[0] + splitValues[1])
                utm_northing = float(splitValues[2] + splitValues[3])
                utm_zone = int(splitValues[5])
                #converting
                mgrs_coordinates = convert.UTMToMGRS(utm_zone,utm_hemisphere,utm_easting,utm_northing)
                cell.value = cell.value.replace(cellValue, mgrs_coordinates)
    #If you wish to have it output to a different path change here
    workbook.save('C:\\Users\\Jochuah Kyle\\Desktop\\Coding Stuff\\UTM to MGRS Python Converter\\Converted To MGRS.xlsx')
                

def MGRSToUTMConverter():
    convert = mgrs.MGRS()
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]

        print(f"Changing Sheet: {sheet_name}")

        #Getting all of the from columns and coordinating them
        for row in sheet.iter_rows(min_row=2,min_col=columnLetter, max_col=columnLetter):
            for cell in row:
                
                mgrs_coordinates = cell.value

                UTM_coordinates_raw = convert.MGRSToUTM(mgrs_coordinates)

                zone = UTM_coordinates_raw[0]
                easting = int(UTM_coordinates_raw[2])
                northing = int(UTM_coordinates_raw[3])

                northingStr = str(northing)
                eastingStr = str(easting)
                eastingPart1 = eastingStr[0]
                eastingPart2 = eastingStr[1:6]
                northingPart1 = f"0{northingStr[0:2]}"
                northingPart2 = northingStr[2:6]
                
                formattedUTMCoordinates = f"{eastingPart1} {eastingPart2} {northingPart1} {northingPart2} 0     {zone} 20"
                cell.value = cell.value.replace(mgrs_coordinates,formattedUTMCoordinates)
    workbook.save('C:\\Users\\Jochuah Kyle\\Desktop\\Coding Stuff\\UTM to MGRS Python Converter\\Converted To MGRS.xlsx')




def mainFunction():
    
    print("What would you like to use? \n(1): Manual UTM to MGRS Converter. \n(2): Manual MGRS to UTM converter \n(3): Automated Excel sheet UTM to MGRS Converter \n(4): Automated Excel sheet MGRS to UTM converter")
    conversionSelector = int(input("Select an option: "))
    if conversionSelector == 1:
        userInputConverterUTMToMGRS()
    elif conversionSelector == 2:
        userInputConverterMGRSToUTM()
    elif conversionSelector == 3:
        UTMToMGRSConverter()
    elif conversionSelector == 4:
        MGRSToUTMConverter()
    else:
        print("INVALID INPUT")

mainFunction()
